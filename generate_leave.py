import io
import openpyxl
from openpyxl.styles import Alignment

TEMPLATE_PATH = "艾迪英特請假單_template.xlsx"

LEAVE_CELLS = {
    "事假":      ("D8",  "D29"),
    "病假":      ("F8",  "F29"),
    "特休":      ("H8",  "H29"),
    "產/婚/喪假": ("J8",  "J29"),
    "其他":      ("M8",  "M29"),
}

def _set(ws, addr, value, halign="center"):
    cell = ws[addr]
    cell.value = value
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal=halign,
        vertical=existing.vertical or "center",
        wrap_text=existing.wrap_text,
    )

def generate_leave_xlsx(
    apply_year, apply_month, apply_day,
    applicant, proxy, leave_type, reason,
    start_year, start_month, start_day, start_hour, start_minute,
    end_year, end_month, end_day, end_hour, end_minute,
    total_days,
) -> bytes:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["假單"]
    ws.print_options.horizontalCentered = True

    def fill(addr, val): _set(ws, addr, val)

    for copy_offset, rows in [
        (0,  dict(date_y="I4",  date_m="K4",  date_d="M4",
                  applicant="D5", proxy="K5",
                  reason="D9",
                  sy="C11", sm="E11", sd="G11", sh="I11", smi="K11",
                  ey="C12", em="E12", ed="G12", eh="I12", emi="K12",
                  days="M11")),
        (21, dict(date_y="I25", date_m="K25", date_d="M25",
                  applicant="D26", proxy="K26",
                  reason="D30",
                  sy="C32", sm="E32", sd="G32", sh="I32", smi="K32",
                  ey="C33", em="E33", ed="G33", eh="I33", emi="K33",
                  days="M32")),
    ]:
        fill(rows["date_y"], apply_year)
        fill(rows["date_m"], apply_month)
        fill(rows["date_d"], apply_day)
        fill(rows["applicant"], applicant)
        fill(rows["proxy"], proxy)
        fill(rows["reason"], reason)
        fill(rows["sy"], start_year)
        fill(rows["sm"], start_month)
        fill(rows["sd"], start_day)
        fill(rows["sh"], start_hour)
        fill(rows["smi"], f"{start_minute:02d}")
        fill(rows["ey"], end_year)
        fill(rows["em"], end_month)
        fill(rows["ed"], end_day)
        fill(rows["eh"], end_hour)
        fill(rows["emi"], f"{end_minute:02d}")
        fill(rows["days"], total_days)

    # Clear all checkbox cells then mark selected
    all_ck = [c for pair in LEAVE_CELLS.values() for c in pair]
    for addr in all_ck:
        ws[addr].value = ""
    if leave_type in LEAVE_CELLS:
        c1, c2 = LEAVE_CELLS[leave_type]
        fill(c1, "○")
        fill(c2, "○")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
