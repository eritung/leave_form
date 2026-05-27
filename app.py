from __future__ import annotations

import io
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

TEMPLATE_PATH = Path(__file__).with_name("艾迪英特請假單.xlsx")
DEFAULT_SHEET_NAME = "假單"

# 兩聯：公司留存 / 請假者留存
SECTION_ROWS = [
    {"date": 4, "name": 5, "leave_check": 8, "reason": 9, "start": 11, "end": 12, "sign_box": 14},
    {"date": 25, "name": 26, "leave_check": 29, "reason": 30, "start": 32, "end": 33, "sign_box": 35},
]

LEAVE_TYPE_MAP = {
    "事假": "D",
    "病假": "F",
    "特休": "H",
    "產假": "J",
    "婚假": "J",
    "喪假": "J",
    "其他": "M",
}

HOLIDAY_PRESETS = {
    "事假": "私人事務需請假處理",
    "病假": "身體不適，需休養就醫",
    "特休": "年度特休假",
    "產假": "產假申請",
    "婚假": "婚假申請",
    "喪假": "喪假申請",
    "其他": "",
}


def roc_year(d: date) -> int:
    return d.year - 1911


def calculate_leave_days(start_date: date, start_time: time, end_date: date, end_time: time) -> float:
    start_dt = datetime.combine(start_date, start_time)
    end_dt = datetime.combine(end_date, end_time)
    if end_dt <= start_dt:
        return 0.0
    hours = (end_dt - start_dt).total_seconds() / 3600
    return round(hours / 8, 2)


def safe_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def write_centered(ws, cell: str, value: Any, size: int = 11) -> None:
    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[cell].font = Font(name="PMingLiU", size=size)


def write_left(ws, cell: str, value: Any, size: int = 11) -> None:
    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[cell].font = Font(name="PMingLiU", size=size)


def pick_sheet(workbook):
    if DEFAULT_SHEET_NAME in workbook.sheetnames:
        return workbook[DEFAULT_SHEET_NAME]
    return workbook.active


def load_template_workbook(uploaded_template):
    if uploaded_template is not None:
        return load_workbook(io.BytesIO(uploaded_template.getvalue()))
    return load_workbook(TEMPLATE_PATH)


def fill_section(ws, rows: dict[str, int], data: dict) -> None:
    date_row = rows["date"]
    name_row = rows["name"]
    leave_check_row = rows["leave_check"]
    reason_row = rows["reason"]
    start_row = rows["start"]
    end_row = rows["end"]
    sign_box_row = rows["sign_box"]

    apply_date = data["apply_date"]
    write_centered(ws, f"I{date_row}", roc_year(apply_date))
    write_centered(ws, f"K{date_row}", apply_date.month)
    write_centered(ws, f"M{date_row}", apply_date.day)

    write_centered(ws, f"D{name_row}", data["applicant"])
    write_centered(ws, f"K{name_row}", data["proxy"])

    for col in ["D", "F", "H", "J", "M"]:
        write_centered(ws, f"{col}{leave_check_row}", "")

    selected_type = data["leave_type"]
    selected_col = LEAVE_TYPE_MAP[selected_type]
    if selected_type in ["產假", "婚假", "喪假"]:
        check_text = f"✓ {selected_type}"
    elif selected_type == "其他":
        note = data.get("other_leave_note", "").strip()
        check_text = f"✓ {note}" if note else "✓"
    else:
        check_text = "✓"
    write_centered(ws, f"{selected_col}{leave_check_row}", check_text)

    write_left(ws, f"D{reason_row}", data["reason"])

    start_date = data["start_date"]
    end_date = data["end_date"]
    start_time = data["start_time"]
    end_time = data["end_time"]

    write_centered(ws, f"C{start_row}", roc_year(start_date))
    write_centered(ws, f"E{start_row}", start_date.month)
    write_centered(ws, f"G{start_row}", start_date.day)
    write_centered(ws, f"I{start_row}", f"{start_time.hour:02d}")
    write_centered(ws, f"K{start_row}", f"{start_time.minute:02d}")
    write_centered(ws, f"M{start_row}", data["leave_days"])

    write_centered(ws, f"C{end_row}", roc_year(end_date))
    write_centered(ws, f"E{end_row}", end_date.month)
    write_centered(ws, f"G{end_row}", end_date.day)
    write_centered(ws, f"I{end_row}", f"{end_time.hour:02d}")
    write_centered(ws, f"K{end_row}", f"{end_time.minute:02d}")

    write_centered(ws, f"C{sign_box_row}", data.get("bu_sign", ""), size=12)
    write_centered(ws, f"G{sign_box_row}", data.get("hr_sign", ""), size=12)
    write_centered(ws, f"K{sign_box_row}", data.get("manager_sign", ""), size=12)


def build_leave_form_xlsx(data: dict, uploaded_template=None) -> bytes:
    wb = load_template_workbook(uploaded_template)
    ws = pick_sheet(wb)

    for rows in SECTION_ROWS:
        fill_section(ws, rows, data)

    ws.print_area = "A1:N39"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


st.set_page_config(page_title="艾迪英特請假單產生器", page_icon="🌸", layout="centered")

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Serif+TC:wght@500;700&family=Noto+Sans+TC:wght@400;500;700&display=swap');

    :root {
        --cream: #fbf6ef;
        --paper: #fffdf8;
        --ink: #3f352b;
        --muted: #8b7f73;
        --sakura: #ead2cf;
        --matcha: #8c9b76;
        --line: #eadfd2;
        --shadow: rgba(91, 71, 49, 0.10);
    }

    .stApp {
        background:
            radial-gradient(circle at top left, rgba(234, 210, 207, 0.45), transparent 30%),
            linear-gradient(180deg, #fbf6ef 0%, #fffaf2 45%, #f8efe4 100%);
        color: var(--ink);
        font-family: 'Noto Sans TC', sans-serif;
    }

    .block-container {
        max-width: 960px;
        padding-top: 2.2rem;
        padding-bottom: 3rem;
    }

    .jp-hero {
        position: relative;
        padding: 2rem 2rem 1.6rem;
        border: 1px solid rgba(206, 184, 161, 0.55);
        border-radius: 28px;
        background: rgba(255, 253, 248, 0.88);
        box-shadow: 0 16px 38px var(--shadow);
        overflow: hidden;
        margin-bottom: 1.2rem;
    }

    .jp-hero::after {
        content: '休';
        position: absolute;
        right: 28px;
        top: -20px;
        font-family: 'Noto Serif TC', serif;
        font-size: 8rem;
        color: rgba(234, 210, 207, 0.32);
        line-height: 1;
    }

    .eyebrow {
        letter-spacing: .18em;
        font-size: .8rem;
        color: var(--matcha);
        font-weight: 700;
        margin-bottom: .35rem;
    }

    .jp-title {
        font-family: 'Noto Serif TC', serif;
        font-size: 2.1rem;
        font-weight: 700;
        line-height: 1.25;
        margin: 0;
        color: var(--ink);
    }

    .jp-subtitle {
        margin-top: .75rem;
        max-width: 680px;
        color: var(--muted);
        line-height: 1.75;
        font-size: .98rem;
    }

    .jp-card {
        padding: 1.25rem 1.35rem;
        border: 1px solid var(--line);
        border-radius: 22px;
        background: rgba(255, 253, 248, 0.92);
        box-shadow: 0 10px 28px rgba(91, 71, 49, 0.075);
        margin: 1rem 0;
    }

    .section-label {
        display: inline-flex;
        align-items: center;
        gap: .45rem;
        padding: .28rem .72rem;
        border-radius: 999px;
        background: #f3e7de;
        color: #66513f;
        font-size: .88rem;
        font-weight: 700;
        margin-bottom: .85rem;
    }

    .soft-note {
        padding: .82rem 1rem;
        border-left: 4px solid var(--matcha);
        border-radius: 14px;
        background: rgba(140, 155, 118, 0.12);
        color: #5f674d;
        line-height: 1.7;
        font-size: .92rem;
    }

    .summary-box {
        border-radius: 18px;
        border: 1px dashed #cdbba7;
        background: rgba(255,255,255,.55);
        padding: 1rem 1.1rem;
        line-height: 1.8;
        color: #5a4b3f;
    }

    div[data-testid="stForm"] {
        border: 0;
        background: transparent;
    }

    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input,
    div[data-testid="stDateInput"] input,
    div[data-testid="stTimeInput"] input,
    textarea {
        border-radius: 14px !important;
        border: 1px solid #dfd1c2 !important;
        background: #fffefb !important;
    }

    div[data-baseweb="select"] > div {
        border-radius: 14px !important;
        border-color: #dfd1c2 !important;
        background: #fffefb !important;
    }

    .stButton > button, .stDownloadButton > button, button[kind="primary"] {
        border-radius: 999px !important;
        border: 0 !important;
        background: linear-gradient(135deg, #8c9b76, #697756) !important;
        color: white !important;
        font-weight: 700 !important;
        box-shadow: 0 10px 22px rgba(105, 119, 86, .24) !important;
        min-height: 3rem;
    }

    .stDownloadButton > button:hover, .stButton > button:hover {
        filter: brightness(1.03);
        transform: translateY(-1px);
    }

    label, .stMarkdown, .stCaptionContainer {
        color: var(--ink) !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="jp-hero">
      <div class="eyebrow">AD2 LEAVE FORM</div>
      <h1 class="jp-title">艾迪英特請假單產生器</h1>
      <div class="jp-subtitle">
        以原始 Excel 假單為模板，填完表單後自動輸出公司留存與請假者留存兩聯。也可以上傳新版假單模板，沿用同一套填寫流程。
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.expander("🌸 模板設定｜可直接上傳原本的 xlsx", expanded=True):
    st.markdown(
        """
        <div class="soft-note">
        不上傳也可以，系統會使用內建模板。若你手上有新版或原本的假單 xlsx，可以在這裡上傳；只要主要欄位位置沒有大幅移動，就能直接套用。
        </div>
        """,
        unsafe_allow_html=True,
    )
    uploaded_template = st.file_uploader(
        "上傳假單 Excel 模板（選填）",
        type=["xlsx"],
        help="建議使用與原始假單相同版型。若欄位位置改很多，需要同步調整座標對應。",
    )

with st.form("leave_form"):
    st.markdown('<div class="jp-card"><div class="section-label">一、基本資料</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        apply_date = st.date_input("申請日期", value=date.today())
        applicant = st.text_input("請假人", placeholder="例如：王小明")
    with col2:
        proxy = st.text_input("代理人", placeholder="可留空")
        leave_type = st.selectbox("假別", ["事假", "病假", "特休", "產假", "婚假", "喪假", "其他"])

    other_leave_note = ""
    if leave_type == "其他":
        other_leave_note = st.text_input("其他假別說明", placeholder="例如：公假、補休、生理假")

    preset_reason = HOLIDAY_PRESETS.get(leave_type, "")
    reason = st.text_area("事由", value=preset_reason, placeholder="請簡述請假原因", height=95)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="jp-card"><div class="section-label">二、請假時間</div>', unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        start_date = st.date_input("開始日期", value=date.today())
    with col2:
        start_time = st.time_input("開始時間", value=time(9, 0))
    with col3:
        end_date = st.date_input("結束日期", value=date.today())
    with col4:
        end_time = st.time_input("結束時間", value=time(18, 0))

    suggested_days = calculate_leave_days(start_date, start_time, end_date, end_time)
    leave_days = st.number_input(
        "請假天數",
        min_value=0.0,
        max_value=90.0,
        value=suggested_days if suggested_days > 0 else 1.0,
        step=0.5,
        help="系統先用起訖時間除以 8 小時估算，仍可手動修改。",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="jp-card"><div class="section-label">三、簽核欄位｜可留空</div>', unsafe_allow_html=True)
    st.caption("若公司流程仍需紙本手簽，這三欄可以不填，輸出的假單會保留空白欄位。")
    col1, col2, col3 = st.columns(3)
    with col1:
        bu_sign = st.text_input("執行長 / BU主管")
    with col2:
        hr_sign = st.text_input("人事部")
    with col3:
        manager_sign = st.text_input("直屬主管")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="jp-card"><div class="section-label">四、輸出前確認</div>', unsafe_allow_html=True)
    st.markdown(
        f"""
        <div class="summary-box">
        <b>請假人：</b>{safe_text(applicant) or '尚未填寫'}<br>
        <b>假別：</b>{leave_type}{'／' + safe_text(other_leave_note) if leave_type == '其他' and safe_text(other_leave_note) else ''}<br>
        <b>期間：</b>{start_date.strftime('%Y/%m/%d')} {start_time.strftime('%H:%M')} ～ {end_date.strftime('%Y/%m/%d')} {end_time.strftime('%H:%M')}<br>
        <b>天數：</b>{leave_days} 天<br>
        <b>模板：</b>{uploaded_template.name if uploaded_template else '使用內建模板'}
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)

    submitted = st.form_submit_button("產生請假單 xlsx", use_container_width=True)

if submitted:
    errors = []
    if not applicant.strip():
        errors.append("請填寫請假人。")
    if not reason.strip():
        errors.append("請填寫事由。")
    if leave_type == "其他" and not other_leave_note.strip():
        errors.append("假別選擇其他時，請填寫其他假別說明。")
    if datetime.combine(end_date, end_time) <= datetime.combine(start_date, start_time):
        errors.append("結束時間需晚於開始時間。")

    if errors:
        for error in errors:
            st.error(error)
    else:
        data = {
            "apply_date": apply_date,
            "applicant": applicant.strip(),
            "proxy": proxy.strip(),
            "leave_type": leave_type,
            "other_leave_note": other_leave_note.strip(),
            "reason": reason.strip(),
            "start_date": start_date,
            "start_time": start_time,
            "end_date": end_date,
            "end_time": end_time,
            "leave_days": leave_days,
            "bu_sign": bu_sign.strip(),
            "hr_sign": hr_sign.strip(),
            "manager_sign": manager_sign.strip(),
        }
        try:
            xlsx_bytes = build_leave_form_xlsx(data, uploaded_template)
            filename = f"艾迪英特請假單_{applicant.strip()}_{apply_date.strftime('%Y%m%d')}.xlsx"
            st.success("請假單已產生，可以下載囉。")
            st.download_button(
                label="下載 xlsx",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as exc:
            st.error("模板讀取或輸出失敗，請確認上傳的是 .xlsx，且版型與原假單相近。")
            st.exception(exc)
